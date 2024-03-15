# import modules
import csv
import pandas as pd
import os
import openpyxl
import string
import easygui
from datetime import date

ResultDate= easygui.enterbox("Enter Based/Proposed, Case Ver. to be printed as Title:")
if ResultDate =="":
    ResultDate = f"Default:{date.today()}"

working_folder = __file__
working_folder = working_folder.rpartition('\\')[0]


# ###use this if u working in maindirectory with folder "Run"
# vtk_list = []
# run_folder = working_folder
# for file in os.listdir(run_folder):
#     if file.endswith(".vtk"):
#         print(os.path.join(run_folder, file))
#         vtk_list.append(os.path.join(file))
#
# if not vtk_list:    #go find other files in run folder if no vtk found in main folder
#     run_folder = working_folder + '\Run'
#     for file in os.listdir(run_folder):
#         if file.endswith(".vtk"):
#             print(os.path.join(run_folder, file))
#             vtk_list.append(os.path.join(file))
# print(vtk_list)

# ###use this if working in subfolder and theres 11vtks
vtk_list = ['run001.vtk', 'run002.vtk', 'run003.vtk', 'run004.vtk','run005.vtk', 'run006.vtk','run007.vtk', 'run008.vtk','run009.vtk', 'run010.vtk','run011.vtk']    #for development purpose

columnsicare= ['AP ID', 'Vel1_Magnitude']
vel1_mag_list = []
apid= []
for vtk in vtk_list:
    df = pd.read_csv(working_folder + f'\\{vtk}_AP_results.csv')
    # print(df)
    if len(apid)==0:
        apid = df['AP ID'].tolist()
    temp_vel1_mag= df['Vel1_Magnitude'].tolist()
    vel1_mag_list.append(temp_vel1_mag)

vel1_mag_list.insert(0, apid)
print(vel1_mag_list)
df = pd.DataFrame(data=vel1_mag_list)
df = df.transpose()
print(df)

writer = pd.ExcelWriter('summary.xlsx', engine='xlsxwriter')
df.to_excel(writer, startcol=1, startrow=3, index= False, header= False)
writer.save()


runs = [x.rpartition('.')[0] for x in vtk_list]
print(runs)
angles = [22.5*int(y.partition('0')[2]) for y in runs]
print(angles)



wb = openpyxl.load_workbook(filename= 'summary.xlsx')
ws = wb['Sheet1']

ws['A1']= ResultDate
for col_num, data in enumerate(runs):
    ws.cell(row=3, column=col_num+3).value= data
    ws.cell(row=3, column=col_num+15).value= data
for col_num, data in enumerate(angles):
    ws.cell(row=2, column=col_num+3).value= data
    ws.cell(row=2, column=col_num+15).value= data


ws['O1']= 6.5

colrange = list(string.ascii_uppercase[14:25])   #O:Y
# print(colrange)
rowrange = range(4,len(df)+4)
# print(rowrange)
colrangeinnumber = [ord(x)-64 for x in colrange]    #not used but incorporated;
# print(colrangeinnumber)

yellowFill = openpyxl.styles.PatternFill(start_color='00FFFF00',
                   end_color='00FFFF00',
                   fill_type='solid')
for row in rowrange:

    for column in colrange:

        ws.cell(row=row, column=ord(column)-64).value = f"={chr(ord(column)-12)}{row}/$O$1"
        ws.cell(row=row, column=ord(column)-64).number_format = '#,##0.00'
        ws.cell(row=row, column=ord(column)-64).fill = yellowFill
        # print(f"={chr(ord(column)-12)}{row}/$O$1")
        # print(ws.cell(row=row, column=ord(column)-64))

wb.save(filename='summary.xlsx')
print("completed")